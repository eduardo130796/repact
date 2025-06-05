import streamlit as st
from openpyxl import load_workbook

def ler_linhas_a_f(arquivo, aba):
    wb = load_workbook(filename=arquivo, data_only=True)
    ws = wb[aba]
    linhas = []
    for row in ws.iter_rows(min_row=1, max_col=6):
        campo_cell = row[0]
        if campo_cell.value is None or str(campo_cell.value).strip() == "":
            continue

        linha = [str(campo_cell.value).strip()]
        for cell in row[1:]:
            valor = cell.value
            formato = cell.number_format

            if formato and "%" in formato:
                tipo = "percentual"
            elif formato and ("R$" in formato or "[$" in formato or "¤" in formato or "0.00" in formato):
                tipo = "reais"
            else:
                tipo = "numero"

            linha.append((valor, tipo))
        linhas.append(linha)
    return linhas

def comparar_linhas(linhas_antigo, linhas_novo):
    max_len = max(len(linhas_antigo), len(linhas_novo))
    comparacao = []
    for i in range(max_len):
        linha_a = linhas_antigo[i] if i < len(linhas_antigo) else [None] + [(None, "numero")] * 5
        linha_b = linhas_novo[i] if i < len(linhas_novo) else [None] + [(None, "numero")] * 5

        campo_a = linha_a[0]
        campo_b = linha_b[0]
        campo = campo_a if campo_a == campo_b else f"{campo_a or ''} / {campo_b or ''}"

        comparacao.append({
            'campo': campo,
            'antigo_B': linha_a[1][0],
            'novo_B': linha_b[1][0],
            'tipo_B': linha_a[1][1],
            'antigo_C': linha_a[2][0],
            'novo_C': linha_b[2][0],
            'tipo_C': linha_a[2][1],
            'antigo_D': linha_a[3][0],
            'novo_D': linha_b[3][0],
            'tipo_D': linha_a[3][1],
            'antigo_E': linha_a[4][0],
            'novo_E': linha_b[4][0],
            'tipo_E': linha_a[4][1],
            'antigo_F': linha_a[5][0],
            'novo_F': linha_b[5][0],
            'tipo_F': linha_a[5][1],
        })
    return comparacao

def formatar_tipo(val, tipo):
    if val in [None, ""]:
        return "-"
    try:
        val_f = float(val)
        if tipo == "percentual":
            return f"{val_f:.2%}"
        elif tipo == "reais":
            return f"R$ {val_f:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
        else:
            return f"{val_f:.2f}".replace(".", ",")
    except:
        return str(val)
st.set_page_config(layout="centered")  # centraliza e limita largura da página
st.title("Comparação Visual - PCF")
col1, col2 = st.columns(2)
with col1:
    arquivo_antigo = st.file_uploader("Planilha Antiga", type="xlsx")
with col2:    
    arquivo_novo = st.file_uploader("Planilha Nova", type="xlsx")

if arquivo_antigo and arquivo_novo:
    wb_antigo = load_workbook(filename=arquivo_antigo, data_only=True)
    wb_novo = load_workbook(filename=arquivo_novo, data_only=True)
    with col1:
        aba_antiga = st.selectbox("Selecione a aba da planilha ANTIGA", wb_antigo.sheetnames)
    with col2:    
        aba_nova = st.selectbox("Selecione a aba da planilha NOVA", wb_novo.sheetnames)

    if aba_antiga and aba_nova:
        linhas_antigo = ler_linhas_a_f(arquivo_antigo, aba_antiga)
        linhas_novo = ler_linhas_a_f(arquivo_novo, aba_nova)

        comparacao = comparar_linhas(linhas_antigo, linhas_novo)


        # Função para checar se tem diferença relevante nas colunas C-F
        def tem_diferenca(item):
            for col in ['C', 'D', 'E', 'F']:
                val_antigo = item.get(f'antigo_{col}')
                val_novo = item.get(f'novo_{col}')
                if (val_antigo or val_novo) and val_antigo != val_novo:
                    return True
            return False

        for item in comparacao:
            if not tem_diferenca(item):
                continue

            mostrar = False
            for col in ['C', 'D', 'E', 'F']:
                val_antigo = item.get(f'antigo_{col}')
                val_novo = item.get(f'novo_{col}')
                if (val_antigo or val_novo) and val_antigo != val_novo:
                    mostrar = True
                    break
            if not mostrar:
                continue

            era_zerado_e_virou_valor = any(
                (item.get(f'antigo_{col}') in [None, "", 0, "0", "0.0"]) and
                (item.get(f'novo_{col}') not in [None, "", 0, "0", "0.0"])
                for col in ['C', 'D', 'E', 'F']
            )

            # Header compacto, texto antigo/novo lado a lado só se houve entrada nova
            if era_zerado_e_virou_valor:
                st.markdown(f"""
                    <div style="
                        background:#fff9e6; 
                        border-left:4px solid #fbbc05; 
                        padding:6px 8px; 
                        margin:6px 0 4px 0; 
                        border-radius:5px;
                        font-size:13px;
                        display:flex;
                        justify-content:space-between;
                        gap:10px;
                    ">
                        <div style="flex:1; color:#1a73e8; font-weight:600;">
                            Antigo:<br><span style='color:#333;'>{item['antigo_B'] or '-'}</span>
                        </div>
                        <div style="flex:1; color:#ea4335; font-weight:600;">
                            Novo:<br><span style='color:#333;'>{item['novo_B'] or '-'}</span>
                        </div>
                    </div>
                """, unsafe_allow_html=True)
            else:
                st.markdown(f"""
                    <div style="
                        background:#f0f0f0; 
                        border-left:4px solid #1a73e8; 
                        padding:6px 10px; 
                        margin:6px 0 4px 0; 
                        border-radius:5px;
                        font-weight:600;
                        font-size:13px;
                        color:#333;
                    ">
                        {item['campo']} - {item['antigo_B'] or '-'}
                    </div>
                """, unsafe_allow_html=True)
            # Acordeão para compactar visual
            
            st.markdown("""
                <style>
                    .grid-header {
                        display: grid; 
                        grid-template-columns: 1.5fr 1.5fr 1fr; 
                        font-weight: 600; 
                        font-size: 13px; 
                        color: #2c3e50; 
                        padding-bottom: 2px; 
                        border-bottom: 1px solid #bbb;
                        margin-bottom: 6px;
                    }
                    .grid-row {
                        display: grid; 
                        grid-template-columns: 1.5fr 1.5fr 1fr;
                        padding: 2px 0;
                        font-size: 12px;
                        border-bottom: 1px solid #eee;
                    }
                    .antigo { color: #1a73e8; }
                    .novo { color: #f9ab00; }
                    .diferenca { color: #d93025; font-weight: 700; }
                    .novo-destaque { background-color: #e6f4ea; font-weight: 600; border-radius: 4px; padding: 1px 4px; display: inline-block;}
                </style>
            """, unsafe_allow_html=True)

            st.markdown(f"""
                <div class="grid-header">
                    <div>Antigo</div><div>Novo</div><div>Diferença</div>
                </div>
            """, unsafe_allow_html=True)

            for col in ['C', 'D', 'E', 'F']:
                val_antigo = item.get(f'antigo_{col}')
                val_novo = item.get(f'novo_{col}')
                tipo = item.get(f'tipo_{col}')

                if (val_antigo in [None, ""]) and (val_novo in [None, ""]):
                    continue

                try:
                    val1 = float(val_antigo)
                    val2 = float(val_novo)
                    dif = val2 - val1
                    pct = (dif / val1) * 100 if val1 != 0 else None
                    cor = "#34a853" if dif > 0 else "#ea4335" if dif < 0 else "#333"
                    dif_formatado = f"<span class='diferenca'>R$ {dif:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".") + "</span>"
                    if pct is not None:
                        dif_formatado += f" <span style='font-weight:400;'>{pct:+.1f}%</span>"
                except:
                    dif_formatado = "<span style='color:#888;'>–</span>"

                if val_antigo == val_novo:
                    dif_formatado = "<span style='color:#aaa;'>–</span>"

                destaque_estilo = ""
                destaque_text = ""
                if (val_antigo in [None, "", 0, "0", "0.0"]) and (val_novo not in [None, "", 0, "0", "0.0"]):
                    destaque_estilo = "background-color: #e6f4ea; font-weight: 600; border-radius: 4px; padding: 1px 4px; display: inline-block;"
                    destaque_text = " 🆕"

                st.markdown(f"""
                    <div class="grid-row">
                        <div class="antigo">{formatar_tipo(val_antigo, tipo)}</div>
                        <div class="novo" style="{destaque_estilo}">{formatar_tipo(val_novo, tipo)}{destaque_text}</div>
                        <div>{dif_formatado}</div>
                    </div>
                """, unsafe_allow_html=True)